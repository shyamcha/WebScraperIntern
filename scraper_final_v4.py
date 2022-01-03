import tkinter as tk
from tkinter import *
from tkinter import filedialog
from pathlib import Path
import requests
import os
import threading
from bs4 import BeautifulSoup
import numpy as np
from time import sleep
from random import randint
#from selenium import webdriver
import requests
import urllib
from urllib.request import Request, urlopen
import re
import csv
from collections import Counter
from string import punctuation
import openpyxl
from pathlib import Path
from multiprocessing import Pool
import concurrent.futures
import tqdm
import cchardet
import lxml
import certifi
import time
from datetime import datetime
import tkinter as tk
from scrapergui_support import *



###Exception
class ScrapingStopped(Exception):
    pass


####SCRAPER CODE####

#Global variables and constants
MAX_THREADS = 250
MIN_THREADS = 7
is_scraping = False
is_stopping = False
keywords = []
website_email_list = []
minimum_wait = 1
maximum_wait = 4
max_websites = 0
current_website = 0


#Textbox editing function
def printb(info):
    StatusBox.configure(state="normal")
    StatusBox.insert('1.0', info +'\n')
    StatusBox.configure(state="disabled")

#Scraping Functions

def getWebPages(url):
  page=url
  url_collected = []
  urls_final = []
  url_final = []
  final_list = []
  #driver.get(page)
  #soup = BeautifulSoup(driver.page_source, 'html.parser')
  r = requests.get(url, headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/42.0.2311.135 Safari/537.36 Edge/12.246'}, timeout = 300)
  soup = BeautifulSoup(r.content, 'lxml')
  #r = scrapy.http.Request(*args, **kwargs)
  urls = [item.get("href") for item in soup.find_all("a")]
  url_collected += urls
  urls_final.clear()
  url_final.clear()
  final_list.clear()
  urls_final = list(dict.fromkeys(url_collected))
  urls_final = list(filter(None, urls_final))
  url_final = [x for x in urls_final if '@' not in x]
  url_final = [x for x in url_final if 'facebook' not in x]
  url_final = [x for x in url_final if 'google' not in x]
  url_final = [x for x in url_final if 'yahoo' not in x]
  url_final = [x for x in url_final if 'facebook' not in x]
  url_final = [x for x in url_final if 'instagram' not in x]
  url_final = [x for x in url_final if 'twitter' not in x]
  url_final = [x for x in url_final if 'linkedin' not in x]
  url_final = [x for x in url_final if 'youtube' not in x]
  url_final = [x for x in url_final if '.jpg' not in x]
  string = page


  final_list = []
  for s in url_final:
    if 'http'  not in s:
      urltokken = [string + s]
      final_list += urltokken
      #final_list=[string + s for s in url_final]
    else:
      urltokken = [s]
      final_list += urltokken

  return final_list

def getSoup(url):
  
  global is_scraping
  if is_scraping == False:
      raise ScrapingStopped("Scraping was manually stopped")

  print(url)
  minimum = minimum_wait
  maximum = maximum_wait
  try:
    sleep(randint(minimum,maximum))
  except Exception as e:
    print(e)
    sleep(randint(1,4))
  r = requests.get(url, headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/42.0.2311.135 Safari/537.36 Edge/12.246'})
  soup = BeautifulSoup(r.content, 'lxml')
  if is_scraping == False:
      raise ScrapingStopped("Scraping was manually stopped")
  return soup

def getWebWords(soup):
  
  # We get the words within paragrphs
  text_p = (''.join(s.findAll(text=True))for s in soup.findAll('p'))
  c_p = Counter((x.rstrip(punctuation).lower() for y in text_p for x in y.split()))

  # We get the words within divs
  text_div = (''.join(s.findAll(text=True))for s in soup.findAll('div'))
  c_div = Counter((x.rstrip(punctuation).lower() for y in text_div for x in y.split()))

  # We sum the two countesr and get a list with words count from most to less common
  total = c_div + c_p
  list_most_common_words = total.most_common() 
  return list_most_common_words

def searchKeyWords(keyword, list_most_common_words):
  #search_list = [index for (index, a_tuple) in enumerate(list_most_common_words) if a_tuple[0]==keyword]
  search_list = [item for item in list_most_common_words if item[0] == keyword]

  if search_list:
    return search_list[0][1]
  else:
    search_list.append(0)
    return search_list[0]

def searchKeyPhrases(keyphrase, soup):
  keyphrases = re.findall(keyphrase, str(soup).lower())
  count = 0
  for i in keyphrases:
    count += 1
  return count

def scraper(website):
  global current_website
  global max_websites
  global website_email_list
  keyword_occurance = []
  emails = set()
  for keyword in keywords:
    keyword_occurance.append(0)
  
  try:
    page_list = getWebPages(website)
    threads = min(MAX_THREADS, len(page_list))
    with concurrent.futures.ThreadPoolExecutor(max_workers=threads) as executor:
      for result in executor.map(getSoup, page_list):
        soup = result
        word_list = getWebWords(soup)

        #Search Keywords
        for keyword in keywords:
          if " " in keyword:
            #count = searchKeyWords(keyword, word_list)
            count = searchKeyPhrases(keyword.lower(), soup)
            keyword_occurance[keywords.index(keyword)] += count
          else:
            count = searchKeyWords(keyword.lower(), word_list)
            keyword_occurance[keywords.index(keyword)] += count
        
        #Gather emails
        #new_emails = set(re.findall(r"[a-z0-9\.\-+_]+@[a-z0-9\.\-+_]+\.com", str(soup), re.I))
        new_emails = set(re.findall(r"[a-z0-9\.\-+_]+@[a-z0-9\.\-+_]+\.['com','net','biz','org','ca']+", str(soup), re.I))
        emails.update(new_emails)

        
    
    grouped = (website, keyword_occurance)
    website_email_list.append([website, emails])
    current_website += 1
    printb("(" + str(current_website) + "/" + str(max_websites) + ")" + " Finished scraping " + website +
     " Time elapsed: " + str(time.thread_time()) + " seconds")

  except Exception as e:
    grouped = (website, ["Could not connect to website: " + str(e)])
    website_email_list.append(grouped)
    current_website += 1
    printb("(" + str(current_website) + "/" + str(max_websites) + ")" + " Could not scrape " 
    + website + " due to following error: " + str(e))
    
  
  return grouped 
  #return keyword_occurance

#Input, Output, and Control Functions

def read_inputs(webpath,keypath):
  #Open excel files with website lists
  website_file = Path(webpath)
  keyword_file = Path(keypath)
  
  wb_obj = openpyxl.load_workbook(website_file) 
  
  # Read the active sheet:
  sheet = wb_obj.active
  
  #Get website names
  websites = []
  for i in range(1, sheet.max_row+1):
      for j in range(1, sheet.max_column+1):
          cell_obj = sheet.cell(row=i, column=j)
          websites.append(cell_obj.value)
  
  
  websites = list(filter(None, websites))
  
  
  for website in websites:
    if not website.endswith('/'):
      webname = website + '/'
    else:
      webname = website
    if 'http' not in website:
      webname = 'http://' + webname
    websites[websites.index(website)] = webname
  
  websites = [i for n, i in enumerate(websites) if i not in websites[:n]]
  wb_obj.close()
  
  #Get keywords
  wb_obj = openpyxl.load_workbook(keyword_file) 
  
  # Read the active sheet:
  sheet = wb_obj.active
  
  #Get website names
  keywords = []
  for i in range(1, sheet.max_row+1):
      for j in range(1, sheet.max_column+1):
          cell_obj = sheet.cell(row=i, column=j)
          keywords.append(cell_obj.value)
  
  keywords= list(filter(None, keywords))
  keywords = [i for n, i in enumerate(keywords) if i not in keywords[:n]]
  
  print(keywords)
  print(websites)
  wb_obj.close()
  return websites, keywords

def write_output(websites, website_keyword_count, outpath):

  output_keywords = []
  output_list = []
  output_keywords = ["website url"] + keywords
  print(output_keywords)


  # Call a Workbook() function of openpyxl 
  # to create a new blank Workbook object
  wb = openpyxl.Workbook()
    
  # Get workbook active sheet  
  # from the active attribute
  sheet = wb.active
  sheet.title = "Results"

  #write headline
  current_row = 1
  current_column = 1
  cell = sheet.cell(row = current_row, column = current_column)
  for item in output_keywords:
      cell = sheet.cell(row = current_row, column = current_column)
      cell.value = item
      current_column += 1

  for website in websites:
      current_row += 1
      current_column = 1
      cell = sheet.cell(row = current_row, column = current_column)
      cell.value = website
      for item in website_keyword_count[websites.index(website)][1]:
          current_column += 1
          cell = sheet.cell(row = current_row, column = current_column)
          cell.value = item
      #output_list = [website] + website_keyword_count[websites.index(website)][1]
      #writer.writerow(output_list)

  #Do the same thing but with emails
  wb.create_sheet('Emails')
  wb.active = wb['Emails']
  sheet = wb.active
  current_row = 1
  current_column = 1
  cell = sheet.cell(row = current_row, column = current_column)
  cell.value = 'website_url'

  email_dictionary = dict(website_email_list)
  for website in websites:
      current_row += 1
      current_column = 1
      cell = sheet.cell(row = current_row, column = current_column)
      cell.value = website
      for item in email_dictionary[website]:
          current_column += 1
          cell = sheet.cell(row = current_row, column = current_column)
          cell.value = item

  wb.active = wb['Results']


  # datetime object containing current date and time
  now = datetime.now()
  print("now =", now)
  # dd/mm/YY H:M:S
  dt_string = now.strftime("%d-%m-%Y %H%M%S")
  print("date and time =", dt_string)
  output_name = "output " + dt_string +".xlsx"

  if outpath == '':
    wb.save(output_name)
    printb("Output saved in current directory as " + output_name)
  else:
    wb.save(outpath + '/' + output_name)
    printb("Output saved to " + outpath + '/' + output_name)
  
  

  wb.close()

def main_function(webpath, keypath, outpath, minimum, maximum):
    global is_scraping
    global is_stopping
    global keywords
    global minimum_wait
    global maximum_wait
    global max_websites
    
    is_scraping = True
    start = time.time()
    try:
      minimum_wait = int(minimum)
      maximum_wait = int(maximum)
    except:
      minimum_wait = 1
      maximum_wait = 4
    
    websites =[]
    keywords =[]
    websites, keywords = read_inputs(webpath, keypath)
    website_keyword_count = []
    max_websites = len(websites)
    threads = min(MIN_THREADS, len(websites))
    with concurrent.futures.ThreadPoolExecutor(max_workers=threads) as executor:
        for result in executor.map(scraper, websites):
            website_keyword_count.append(result)
    printb("Finished scraping list ") 
    write_output(websites, website_keyword_count, outpath)
    end = time.time()
    printb(f"Runtime of the scrape is {end - start}")
    is_scraping = False
    is_stopping = False


####GUI

#to locate file's location
#important in importing HBM Logo and setting a default output location
global prog_location
prog_call = sys.argv[0]
prog_location = os.path.split(prog_call)[0]


#GUI functions

def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
        
    return os.path.join(base_path, relative_path)

def getWebsiteListPath(): 
    folder_path = filedialog.askopenfilename(filetypes=[("Excel files","*.xlsx")],initialdir = os.getcwd())
    WebsiteBox.delete(0, "end")
    WebsiteBox.insert(0, folder_path)

def getKeywordListPath(): 
    folder_path = filedialog.askopenfilename(filetypes=[("Excel files","*.xlsx")],initialdir = os.getcwd())
    KeywordBox.delete(0, "end")
    KeywordBox.insert(0, folder_path)   

def getOutputPath(): 

    folder_path = filedialog.askdirectory()
    OutputBox.delete(0, "end")
    OutputBox.insert(0, folder_path)  

def clearOutput():
   if is_scraping == False:
      WebsiteBox.delete(0, "end")
      KeywordBox.delete(0, "end")
      OutputBox.delete(0, "end")
      MaxEntry.delete(0, "end")
      MinEntry.delete(0, "end")
      StatusBox.configure(state="normal")
      StatusBox.delete(1.0, "end")
      StatusBox.configure(state="disabled")
      StatusFrame.configure(text='''Status''')

def on_closing():
    window.destroy()
    
def start_program():
    global is_scraping
    global current_website
    try:
        if is_scraping == False & is_stopping == False:
            printb("Scraping list....")
            current_website = 0
            threading.Thread(target = main_function, args = (WebsiteBox.get(), KeywordBox.get(), OutputBox.get(), MinEntry.get(), MaxEntry.get())).start()
        else:
            printb('Scrape already in progress')
    except Exception as e:
        printb(e)
        is_scraping = False

def stop_program():
  global is_scraping
  global is_stopping 
  if is_scraping == False:
    printb("No scrape in progress")
  else:
    is_scraping = False
    printb("Stopping scraping...")
    raise ScrapingStopped

  

#main window

window= tk.Tk()
window.minsize(475, 600)
window.maxsize(475, 845)
window.title("Keyword Scraper")
window.resizable(1,  1)

#Logo
LogoLabel = tk.Label(window)
photo_location = os.path.join(prog_location,resource_path("logo.png"))
#photo_location = Path("logo.png")
global _img0
_img0 = tk.PhotoImage(file=photo_location)
LogoLabel.configure(image=_img0)
LogoLabel.place(relx=0.09, rely=0.014, height=101, width=160)

#Title
TitleLabel = tk.Label(window)
TitleLabel.configure(font="-family {Segoe UI} -size 16")
TitleLabel.configure(text='''Keyword Scraper''')
TitleLabel.place(relx=0.385, rely=0.043, height=51, width=224)

#Website loading widgets
WebsiteLabel = tk.Label(window)
WebsiteLabel.place(relx=0.023, rely=0.166, height=22, width=125)
WebsiteLabel.configure(text='''Website List''')

WebsiteBox = tk.Entry(window)
WebsiteBox.place(relx=0.274, rely=0.166, height=20, relwidth=0.575)

WebsiteButton = tk.Button(window)
WebsiteButton.place(relx=0.844, rely=0.166, height=24, width=57)
WebsiteButton.configure(text='''Browse''')
WebsiteButton.configure(command= getWebsiteListPath)

#Keyword Loading widgets
KeywordLabel = tk.Label()
KeywordLabel.place(relx=0.023, rely=0.241, height=22, width=125)
KeywordLabel.configure(text='''Keyword List''')

KeywordBox = tk.Entry()
KeywordBox.place(relx=0.274, rely=0.241, height=20, relwidth=0.575)

KeywordButton = tk.Button()
KeywordButton.place(relx=0.844, rely=0.242, height=24, width=57)
KeywordButton.configure(text='''Browse''')
KeywordButton.configure(command= getKeywordListPath)

#Outputbox 
OutputLabel = tk.Label()
OutputLabel.place(relx=0.0, rely=0.318, height=22, width=125)
OutputLabel.configure(text='''Output Location''')

OutputBox = tk.Entry()
OutputBox.place(relx=0.274, rely=0.318, height=20, relwidth=0.575)
      
OutputButton = tk.Button()
OutputButton.place(relx=0.844, rely=0.318, height=24, width=57)
OutputButton.configure(text='''Browse''')
OutputButton.configure(command=getOutputPath)

#Delay controls
MinLabel = tk.Label()
MinLabel.place(relx=0.0, rely=0.376, height=22, width=125)
MinLabel.configure(text='''Minimum Delay''')

MaxLabel = tk.Label()
MaxLabel.place(relx=0.0, rely=0.434, height=22, width=125)
MaxLabel.configure(text='''Maximum Delay''')
    
MinEntry = tk.Entry()
MinEntry.place(relx=0.274, rely=0.376, height=20, relwidth=0.054)

MaxEntry = tk.Entry()
MaxEntry.place(relx=0.274, rely=0.434, height=20, relwidth=0.054)


#Buttons            
ClearButton = tk.Button()
ClearButton.place(relx=0.138, rely=0.507, height=34, width=97)
ClearButton.configure(text = "Clear")
ClearButton.configure(command = clearOutput)

StopButton = tk.Button()
StopButton.place(relx=0.387, rely=0.507, height=34, width=97)
StopButton.configure(text='''Stop''')
StopButton.configure(command = stop_program)

StartButton = tk.Button()
StartButton.place(relx=0.636, rely=0.507, height=34, width=97)
StartButton.configure(text='''Start''')
StartButton.configure(command= start_program)
#StartButton.configure(command= lambda: threading.Thread(target =  
 #   main_function(WebsiteBox.get(), KeywordBox.get(), OutputBox.get(), MinEntry.get(), MaxEntry.get())).start())

#Status Frame
StatusFrame = tk.LabelFrame()
StatusFrame.place(relx=0.023, rely=0.594, relheight=0.326
        , relwidth=0.957)
StatusFrame.configure(text='''Status''')
StatusBox = ScrolledText(StatusFrame)
StatusBox.place(relx=0.024, rely=0.102, relheight=0.85, relwidth=0.955, bordermode='ignore')
StatusBox.configure(state="disabled")
StatusBox.configure(wrap="none")

window.protocol("WM_DELETE_WINDOW", on_closing)

window.mainloop()
window.destroy()

